import os
import sys
import re
import networkx as nx
import matplotlib.pyplot as plt 
import numpy as np 
import pandas as pd
import csv
from openpyxl import load_workbook


ex_dir = os.path.dirname(__file__)
cn_module_dir = os.path.join(ex_dir, '../..')
sys.path.append( cn_module_dir )

from src.visuals import *
from src.utils import *
from src.states import *
from src.generators import *
from src.plotting import *
from src.kms_graphs import *
from src.measures import *
import seaborn as sns
# import fitter
# from fitter import Fitter, get_common_distributions
from src.gen_colors import *
from neuron_positions import *


from connectome import synapses, S, neuron_positions, node_colors, node_labels, CIRCUITS, blue_sky
from new_connectome import new_connections

neurons = list(S.nodes)
# cook_con_file = '../data/Cook_connectome_adj_matrices_extra.xlsx'

# workbook = load_workbook(cook_con_file)
# chem_sheet = 'hermaphrodite chemical'
# gap_jn_sheet = 'herm gap jn symmetric'

# # sheets = workbook.sheetnames
# sheet = workbook[gap_jn_sheet]

# # print(sheets)
# cols = sheet[3]
# rows = sheet['C']

# # col_headers = cols[4:471]
# # row_headers = rows[3:455]

# # rows_range = range(4, 470)
# # cols_range = range(3, 469)
# chem_cols_range = range(3, 471)
# chem_rows_range = range(4, 472)

# print(sheet[471][2].value)
# print(cols[455].value)

# syn_list = []

# # gap_jn_list = []

# for i in chem_rows_range:
#     n1 = sheet[i][2].value
#     for j in chem_cols_range:
#         n2 = cols[j].value
#         cell = sheet[i][j].value
#         if cell != None:
#             syn_list += [(n1, n2, cell),]



# len(syn_list)

## Save the synapses in csv file
# new_conn_f = open('../data/extrapolated_connectome_ej.csv', 'w', newline='')
# writer = csv.writer(new_conn_f, delimiter=',')
# writer.writerow(['# Gap junctions in the C. elegans'])
# writer.writerow(['# These data include extrapolated connections from the dataset used by Cook et al.'])
# for syn in syn_list:
#     row = [str(syn[0]),str(syn[1]),syn[2],'EJ']
#     writer.writerow(row)

## Reading edges from csv files

chem_syn_file = open('../data/extrapolated_connectome_chem.csv', 'r', newline='')
ej_file = open('../data/extrapolated_connectome_ej.csv', 'r', newline='')


# chem_lines = chem_syn_file.readlines()[2:]
# ej_lines = ej_file.readlines()[2:]
# # com_lines = chem_lines + ej_lines
# # n_lines = [re.sub('[ ]+', ',', line) for line in f.readlines()]
# chem_rows = [line.strip().split(',') for line in chem_syn_file.readlines()[2:]]
# ej_rows = [line.strip().split(',') for line in ej_file.readlines()[2:]]

# com_rows = chem_rows + ej_rows
# # # print(ej_lines[:10])

# # ## Save the new complete connectome in the same file
# new_conn_f = open('../data/extrapolated_connectome.csv', 'w', newline='')
# writer = csv.writer(new_conn_f, delimiter=',')
# writer.writerow(['# Extended chemical synapses and gap junctions in the C. elegans -- Pharynx excluded'])
# writer.writerow(['# These data include extrapolated connections from the dataset used by Cook et al.'])
# for l in com_rows:
#     if (str(l[0]) in neurons) and (str(l[1]) in neurons):
#         writer.writerow([str(l[0]),str(l[1]),l[2],str(l[3])])

## Read the new edges for the somatic connectome: 279 neurons
ext_con_file = open('../data/extrapolated_connectome.csv', 'r', newline='')
conn_rows = [line.strip().split(',') for line in ext_con_file.readlines()[2:]]

ext_connections = []

for r in conn_rows:
    ext_connections += [(r[0], r[1]),] * int(r[2])



## Compare new connectome with the old one

# added_connect = [e for e in list(set(ext_connections)) if e not in new_connections]

# print(len(added_connect))
    
# Study the graph of the new connectome

ES = nx.MultiDiGraph()
ES.add_edges_from(ext_connections)
# #### Comparaison to the updated connectome
# extended_connect = [e for e in  list(set(ext_connections)) if e not in list(set(synapses))]


# eb_c = critical_inverse_temperature(ES) # = 6.012528109900082, T_c = 0.16631938873656565

# print(len(extended_connect))


#### Draw structural circuits
# name = 'Thermotaxis'
# # # for name in CIRCUITS:
# V = CIRCUITS[name]
# K = weighted_structural_subgraph(ES, nodelist=V)
# draw_weighted_digraph(K, pos={n: neuron_positions[n] for n in K.nodes}, node_size=1.6, node_colors={n: node_colors[n] for n in K.nodes}, node_labels={n: node_labels[n] for n in K.nodes}, edgecolor='silver', font_size=7, figsize=(10,6))
# plt.show()
# plt.savefig(f'./results/nets/struc/{name}_circuit.pdf', dpi=300, bbox_inches='tight')
# plt.close()

# KMS subgraphs
# b_factors = [.9, .98, 1.001, 1.01, 1.06, 1.1, 1.2, 1.5, 1.8, 1.9, 2., 2.1, 2.4, 2.6, 2.8, 3., 3.1, 3.2, 3.5, 3.8, 4.] 
# betas = [2.]
# # # b_factor = 1.06
# name = 'Thermotaxis'
# V = CIRCUITS[name]
# for factor in betas:
#     b = factor * eb_c
#     # V = olfactory_thermo_neurons
#     K = kms_weighted_subgraph(ES, b, nodelist=V)

# # # print(list(K.edges(data=True))[0][2]['weight'])
#     draw_weighted_digraph(K, pos={n: neuron_positions[n] for n in K.nodes}, node_size=1.6, node_colors={n: node_colors[n] for n in K.nodes}, node_labels={n: node_labels[n] for n in K.nodes}, edgecolor=blue_sky, font_size=8, figsize=(10,6))
#     # plt.savefig(f'./results/nets/kms_graphs/func_circuits/{name}_KMS-subgraph-{b_factor}xbeta_c.pdf', dpi=300, bbox_inches='tight')
#     # plt.close()
#     plt.show()



###### DEMO 
# from scipy.cluster.hierarchy import fcluster

# # Generate synthetic data
# np.random.seed(42)
# data = np.random.rand(10, 12)
# df = pd.DataFrame(data, columns=[f'Feature_{i}' for i in range(12)])

# # Create the clustermap
# clustergrid = sns.clustermap(df, method='average', metric='euclidean')

# # Extract the linkage matrix
# linkage_matrix = clustergrid.dendrogram_row.linkage

# # Form clusters using the linkage matrix
# # Define the number of clusters or a distance threshold
# n_clusters = 3
# cluster_labels = fcluster(linkage_matrix, n_clusters, criterion='maxclust')

# # Add the cluster labels to the DataFrame for easy access
# df['Cluster'] = cluster_labels

# # Print the resulting clusters
# print(df)

# # Optionally, plot the data points color-coded by their cluster
# import matplotlib.pyplot as plt

# plt.figure(figsize=(10, 8))
# for cluster in range(1, n_clusters + 1):
#     subset = df[df['Cluster'] == cluster]
#     plt.scatter(subset.index, subset.iloc[:, 0], label=f'Cluster {cluster}')
# plt.legend()
# plt.title('Data points color-coded by their cluster')
# plt.xlabel('Sample Index')
# plt.ylabel('Feature_0')
# plt.show()
