import os
import sys
import re
import networkx as nx
import matplotlib.pyplot as plt 
import numpy as np 
import pandas as pd
import csv
from openpyxl import load_workbook


ex_dir = os.path.dirname(__file__)
cn_module_dir = os.path.join(ex_dir, '../..')
sys.path.append( cn_module_dir )

from src.visuals import *
from src.utils import *
from src.states import *
from src.generators import *
from src.plotting import *
from src.kms_graphs import *
from src.measures import *
import seaborn as sns
# import fitter
# from fitter import Fitter, get_common_distributions
from src.gen_colors import *
from neuron_positions import *


from connectome import synapses, S, neuron_positions, node_colors, node_labels, CIRCUITS, blue_sky


neurons = list(S.nodes)
# cook_file = '../data/synapses_list_updated.xlsx'

# workbook = load_workbook(cook_file)
# herm_sheet = 'hermaphrodite'
# # sheets = workbook.sheetnames
# sheet = workbook[herm_sheet]

# pre_syn = sheet['C']
# post_syn = sheet['D']
# type_syn = sheet['E']
# print(pre_syn[1].value)
# print(len(type_syn))
# print(pre_syn[1].value, post_syn[1].value, type_syn[1].value)

# syn_updated = []

# for i, _ in enumerate(pre_syn):
#     conn_type = type_syn[i].value
#     pre = pre_syn[i].value
#     post_values = post_syn[i].value
#     for p in post_values.split(','):
#         syn_updated += [(str(pre), str(p), str(conn_type)),]




# len(syn_list)

## Save the synapses in csv file
# new_conn_f = open('../data/cook/herm_whole_connectome_updated.csv', 'w', newline='')
# writer = csv.writer(new_conn_f, delimiter=',')
# writer.writerow(['# Gap junctions in the C. elegans'])
# writer.writerow(['# From the synapses list dataset used by Cook et al.'])
# for syn in syn_updated:
#     row = [str(syn[0]),str(syn[1]),str(syn[2])]
#     writer.writerow(row)

## Save only the somatic connectome
# new_conn_f = open('../data/cook/herm_whole_connectome_updated.csv', 'r', newline='')

# syn_rows = [line.strip().split(',') for line in new_conn_f.readlines()[2:]]
# somatic_conn_f = open('../data/cook/herm_somatic_connectome_updated.csv', 'w', newline='')

# writer = csv.writer(somatic_conn_f, delimiter=',')
# writer.writerow(['# Somatic Chemical synapses and gap junctions in the C. elegans'])
# writer.writerow(['# From the synapses list dataset used by Cook et al.'])
# for r in syn_rows:
#     n1 = str(r[0])
#     n2 = str(r[1])
#     if (n1 in neurons) and (n2 in neurons):
#         row = [n1,n2,str(r[2])]
#         writer.writerow(row)


### Reading the file and constructing the new connectome network
new_conn_f = open('../data/cook/herm_somatic_connectome_updated.csv', 'r', newline='')

syn_rows = [line.strip().split(',') for line in new_conn_f.readlines()[2:]]

synapses_uptd = []

for r in syn_rows:

    synapses_uptd += [(r[0], r[1]),] 



# ## Compare new connectome with the old one

# added_connect = [e for e in list(set(synapses)) if e not in  list(set(synapses_uptd))]

# print(len(added_connect))
# afd = ['AFDR', 'AFDL']
# AFD_links = [e for e in synapses_uptd if (e[0] in afd) or (e[1] in afd)]
print(len(synapses))    
# Study the graph of the new connectome

# SU = nx.MultiDiGraph()
# # NS.add_nodes_from(neurons)
# SU.add_edges_from(synapses_uptd)


# # print(len(SU.nodes))
# print([n for n in neurons if n not in list(SU.nodes)])

# nb_c = critical_inverse_temperature(NS) # = 4.495504693131899, T_c =  0.2224444346655384